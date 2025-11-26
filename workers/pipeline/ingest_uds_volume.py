"""
HRSA UDS VOLUME INGESTION
Author: Charta Health GTM Data Engineering

Purpose: Extract verified patient volume from HRSA UDS 2024 reports.

Data Source: HRSA Uniform Data System (UDS) reports for Health Centers
Output: Grant-level patient counts for use in volume verification

2024 UDS File Structure (Strict):
- Target Sheet: Table3A (Patient Demographics)
- Key Column: GrantNumber (Grant ID, e.g., "10030")
- Volume Columns:
  * T3a_L39_Ca = Line 39, Column A (Male Patients)
  * T3a_L39_Cb = Line 39, Column B (Female Patients)
  * Total Patients = T3a_L39_Ca + T3a_L39_Cb
- Data Structure:
  * Row 0 (header): Column names
  * Row 1 (skip): Column descriptions
  * Row 2+: Actual data

Fallback: Fuzzy matching for non-standard files
"""

import pandas as pd
import numpy as np
import os
import sys
from pathlib import Path
import openpyxl
from openpyxl import load_workbook

# Configuration - Go up 2 levels from workers/pipeline/ to project root
ROOT = os.path.abspath(os.path.join(os.path.dirname(__file__), "..", ".."))
DATA_RAW = os.path.join(ROOT, "data", "raw")
DATA_CURATED = os.path.join(ROOT, "data", "curated")
DATA_STAGING = os.path.join(DATA_CURATED, "staging")

# Input/Output
HRSA_DIR = os.path.join(DATA_RAW, "hrsa")
OUTPUT_FILE = os.path.join(DATA_STAGING, "stg_uds_volume.csv")

# Target sheet names (priority order)
TARGET_SHEETS = ['Table 3A', 'Table3A', 'Patients', 'Universal', 'Data', 'Sheet1']

# Common column name variations
GRANT_NUMBER_COLS = ['GrantNumber', 'Grant Number', 'grant_number', 'BHCMIS ID', 'BHCMISID', 
                     'bhcmis_id', 'Health Center ID', 'Grant_Number', 'ID', 'Grantee']
PATIENT_COUNT_COLS = ['Total Patients', 'total_patients', 'Total', 'Patients', 'Line 39', 
                      'Total_Patients', 'TotalPatients', 'Patient Count', 'T3a_L39_Ca', 'T3a_L39_Cb']

def find_uds_files():
    """
    Scan HRSA directory for UDS 2024 Excel files.
    Returns list of file paths.
    """
    print(f"🔍 Scanning {HRSA_DIR} for UDS 2024 files...")
    
    if not os.path.exists(HRSA_DIR):
        print(f"   ❌ HRSA directory not found: {HRSA_DIR}")
        return []
    
    uds_files = []
    
    # Walk through directory and subdirectories
    for root, dirs, files in os.walk(HRSA_DIR):
        for file in files:
            # Check for Excel files with 'uds' or '2024' in name
            if file.endswith(('.xlsx', '.xls')):
                file_lower = file.lower()
                if 'uds' in file_lower or '2024' in file_lower:
                    file_path = os.path.join(root, file)
                    uds_files.append(file_path)
    
    print(f"   ✅ Found {len(uds_files)} UDS files")
    for i, f in enumerate(uds_files, 1):
        print(f"      {i}. {os.path.basename(f)}")
    
    return uds_files

def find_column_index(df, target_cols):
    """
    Find the first matching column from a list of possible names.
    Returns column name or None.
    """
    df_cols_lower = {col.lower(): col for col in df.columns if isinstance(col, str)}
    
    for target in target_cols:
        target_lower = target.lower()
        if target_lower in df_cols_lower:
            return df_cols_lower[target_lower]
    
    return None

def extract_from_sheet_strict(file_path, sheet_name='Table3A'):
    """
    Extract grant number and patient count using strict 2024 UDS structure.
    
    2024 UDS Format:
    - Sheet: Table3A
    - ID Column: GrantNumber
    - Male Patients: T3a_L39_Ca (Line 39, Column A)
    - Female Patients: T3a_L39_Cb (Line 39, Column B)
    - Total Patients = T3a_L39_Ca + T3a_L39_Cb
    - Row 0 contains descriptions, data starts at Row 1
    
    Returns DataFrame with columns ['grant_number', 'uds_patient_count']
    """
    try:
        # Read Table3A with header on row 0
        df = pd.read_excel(file_path, sheet_name=sheet_name, header=0)
        
        if df.empty:
            return None
        
        # Check for 2024 UDS column structure
        if 'GrantNumber' in df.columns and 'T3a_L39_Ca' in df.columns and 'T3a_L39_Cb' in df.columns:
            # Drop first row (contains descriptions)
            df = df.iloc[1:].copy()
            
            # Extract grant number and patient count columns
            result = df[['GrantNumber', 'T3a_L39_Ca', 'T3a_L39_Cb']].copy()
            
            # Convert patient count columns to numeric (coerce errors to NaN)
            result['T3a_L39_Ca'] = pd.to_numeric(result['T3a_L39_Ca'], errors='coerce').fillna(0)
            result['T3a_L39_Cb'] = pd.to_numeric(result['T3a_L39_Cb'], errors='coerce').fillna(0)
            
            # Calculate total patients (Male + Female)
            result['uds_patient_count'] = result['T3a_L39_Ca'] + result['T3a_L39_Cb']
            
            # Rename and select final columns
            result = result[['GrantNumber', 'uds_patient_count']].copy()
            result.columns = ['grant_number', 'uds_patient_count']
            
            # Clean grant numbers
            result['grant_number'] = result['grant_number'].astype(str).str.strip()
            result = result[result['grant_number'].notna() & (result['grant_number'] != '') & (result['grant_number'] != 'nan')]
            
            # Filter for valid patient counts (at least 100)
            result = result[result['uds_patient_count'] >= 100]
            
            if len(result) > 0:
                return result
        
    except Exception as e:
        # Sheet might not exist or have different structure
        pass
    
    return None

def extract_from_sheet_fuzzy(file_path, sheet_name):
    """
    Extract grant number and patient count using fuzzy column matching.
    Fallback method when strict extraction fails.
    Returns DataFrame with columns ['grant_number', 'uds_patient_count']
    """
    try:
        # Try reading the sheet
        df = pd.read_excel(file_path, sheet_name=sheet_name, dtype=str)
        
        if df.empty:
            return None
        
        # Find grant number column
        grant_col = find_column_index(df, GRANT_NUMBER_COLS)
        
        # Find patient count column
        patient_col = find_column_index(df, PATIENT_COUNT_COLS)
        
        if grant_col is None or patient_col is None:
            return None
        
        # Extract data
        result = df[[grant_col, patient_col]].copy()
        result.columns = ['grant_number', 'uds_patient_count']
        
        # Clean grant numbers (remove whitespace, convert to string)
        result['grant_number'] = result['grant_number'].astype(str).str.strip()
        result = result[result['grant_number'].notna() & (result['grant_number'] != '') & (result['grant_number'] != 'nan')]
        
        # Clean patient counts (convert to numeric)
        result['uds_patient_count'] = pd.to_numeric(result['uds_patient_count'], errors='coerce')
        result = result[result['uds_patient_count'].notna() & (result['uds_patient_count'] > 0)]
        
        if len(result) > 0:
            return result
        
    except Exception as e:
        # Silently handle errors (sheet might not exist or have different structure)
        pass
    
    return None

def extract_from_file(file_path):
    """
    Extract UDS volume data from a single Excel file.
    Uses strict 2024 UDS format first, then falls back to fuzzy matching.
    Returns DataFrame or None.
    """
    print(f"   📊 Processing: {os.path.basename(file_path)}")
    
    try:
        # Get list of sheet names
        wb = load_workbook(file_path, read_only=True, data_only=True)
        available_sheets = wb.sheetnames
        wb.close()
        
        print(f"      Sheets available: {', '.join(available_sheets[:5])}{' ...' if len(available_sheets) > 5 else ''}")
        
        # STEP 1: Try strict 2024 UDS format (Table3A with BHCMIS ID and Total Patients)
        if 'Table3A' in available_sheets:
            print(f"      Attempting strict extraction from 'Table3A'...")
            result = extract_from_sheet_strict(file_path, 'Table3A')
            if result is not None and len(result) > 0:
                print(f"      ✅ Extracted {len(result):,} records using strict 2024 UDS format")
                return result
            else:
                print(f"      ⚠️  Strict extraction failed, trying fuzzy matching...")
        
        # STEP 2: Fall back to fuzzy matching on target sheets
        for target_sheet in TARGET_SHEETS:
            # Check if sheet exists (case-insensitive)
            matching_sheet = None
            for sheet in available_sheets:
                if sheet.lower() == target_sheet.lower():
                    matching_sheet = sheet
                    break
            
            if matching_sheet:
                result = extract_from_sheet_fuzzy(file_path, matching_sheet)
                if result is not None and len(result) > 0:
                    print(f"      ✅ Extracted {len(result):,} records from sheet '{matching_sheet}' (fuzzy)")
                    return result
        
        # STEP 3: Last resort - try first sheet with fuzzy matching
        if available_sheets:
            result = extract_from_sheet_fuzzy(file_path, available_sheets[0])
            if result is not None and len(result) > 0:
                print(f"      ✅ Extracted {len(result):,} records from sheet '{available_sheets[0]}' (fuzzy)")
                return result
        
        print(f"      ⚠️  No valid data found")
        
    except Exception as e:
        print(f"      ❌ Error: {str(e)}")
    
    return None

def ingest_uds_volume():
    """
    Main ingestion function.
    Scans for UDS files, extracts volume data, and saves to staging.
    """
    print("="*80)
    print(" INGEST HRSA UDS 2024 VOLUME DATA")
    print("="*80)
    
    # 1. Find UDS files
    uds_files = find_uds_files()
    
    if not uds_files:
        print("\n   ❌ No UDS files found. Check data/raw/hrsa/ directory.")
        return
    
    # 2. Extract from each file
    print(f"\n📥 Extracting data from {len(uds_files)} file(s)...")
    
    all_data = []
    
    for file_path in uds_files:
        data = extract_from_file(file_path)
        if data is not None:
            all_data.append(data)
    
    if not all_data:
        print("\n   ❌ No data extracted. Files may have unexpected structure.")
        print("   💡 Tip: Ensure files contain 'Grant Number' and 'Total Patients' columns.")
        return
    
    # 3. Combine all data
    print(f"\n🔗 Combining data from {len(all_data)} successful extraction(s)...")
    combined = pd.concat(all_data, ignore_index=True)
    
    # 4. Deduplicate (keep highest patient count per grant)
    print("   Deduplicating by grant number (keeping max patient count)...")
    combined = combined.sort_values('uds_patient_count', ascending=False)
    combined = combined.drop_duplicates(subset='grant_number', keep='first')
    
    # 5. Validate
    print(f"   Validating {len(combined):,} records...")
    
    # Filter out invalid grant numbers (too short, all zeros, etc.)
    combined = combined[combined['grant_number'].str.len() >= 5]
    combined = combined[~combined['grant_number'].str.match(r'^0+$')]
    
    # Filter for reasonable patient counts (at least 100, max 500k)
    combined = combined[
        (combined['uds_patient_count'] >= 100) & 
        (combined['uds_patient_count'] <= 500000)
    ]
    
    print(f"   ✅ {len(combined):,} valid records after filtering")
    
    # 6. Save
    print(f"\n💾 Saving to {OUTPUT_FILE}...")
    combined.to_csv(OUTPUT_FILE, index=False)
    print(f"   ✅ Saved {len(combined):,} FQHC volume records")
    
    # 7. Report
    print("\n📊 SUMMARY STATISTICS:")
    print(f"   Total Health Centers: {len(combined):,}")
    print(f"   Total Patient Count: {combined['uds_patient_count'].sum():,.0f}")
    print(f"   Avg Patients per HC: {combined['uds_patient_count'].mean():,.0f}")
    print(f"   Median Patients: {combined['uds_patient_count'].median():,.0f}")
    print(f"   Max Patients: {combined['uds_patient_count'].max():,.0f}")
    print(f"   Min Patients: {combined['uds_patient_count'].min():,.0f}")
    
    # Distribution
    print(f"\n📈 Patient Volume Distribution:")
    print(f"   Small (<5,000): {len(combined[combined['uds_patient_count'] < 5000]):,}")
    print(f"   Medium (5k-20k): {len(combined[(combined['uds_patient_count'] >= 5000) & (combined['uds_patient_count'] < 20000)]):,}")
    print(f"   Large (20k-50k): {len(combined[(combined['uds_patient_count'] >= 20000) & (combined['uds_patient_count'] < 50000)]):,}")
    print(f"   Very Large (50k+): {len(combined[combined['uds_patient_count'] >= 50000]):,}")

def main():
    try:
        ingest_uds_volume()
    except Exception as e:
        print(f"\n❌ Fatal Error: {str(e)}")
        import traceback
        traceback.print_exc()
        sys.exit(1)

if __name__ == "__main__":
    main()

